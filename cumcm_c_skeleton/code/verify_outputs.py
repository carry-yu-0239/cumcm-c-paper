"""Validate workbook structure and render the generated paper figures for inspection."""
from pathlib import Path
from shutil import which
from subprocess import run
from tempfile import TemporaryDirectory

from openpyxl import load_workbook
from PIL import Image, ImageChops

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'data' / '\u9644\u4ef6_\u6570\u636e\u9884\u5904\u7406\u7ed3\u679c.xlsx'
REQUIRED = {'\u5904\u7406\u8bf4\u660e', '\u8868\u53551\u6e05\u6d17\u53ca\u989c\u8272\u586b\u8865', '\u8868\u53552\u6e05\u6d17\u8bb0\u5f55', '\u6709\u6548\u539f\u59cb\u6570\u636e', '\u95ed\u5408\u5f52\u4e00\u5316\u6570\u636e', '\u8fd1\u96f6\u66ff\u6362\u6570\u636e', 'CLR\u539f\u59cb\u503c', 'Z-CLR\u6570\u636e', '\u5f02\u5e38\u8bca\u65ad', '\u8fd1\u96f6\u53c2\u6570\u654f\u611f\u6027', 'U3_closed', 'U3_replace', 'U3_CLR', 'U3_ZCLR'}

if not BOOK.is_file():
    raise FileNotFoundError(BOOK)
wb = load_workbook(BOOK, read_only=True, data_only=True)
missing = REQUIRED - set(wb.sheetnames)
if missing:
    raise AssertionError(f'Missing sheets: {sorted(missing)}')
if wb['Z-CLR\u6570\u636e'].max_column != 16:
    raise AssertionError('Z-CLR must have 2 identifiers plus 14 variables.')
if wb['U3_ZCLR'].max_column != 15:
    raise AssertionError('Unknown Z-CLR must have 1 identifier plus 14 variables.')

renderer = which('pdftoppm')
if not renderer:
    raise RuntimeError('pdftoppm is required for reproducible PDF rendering checks.')
# Codex's Windows command shim may point at a non-existent Poppler subpath.
# Prefer the bundled executable when that layout is present.
shim = Path(renderer)
bundled = shim.parents[2] / 'native' / 'poppler' / 'Library' / 'bin' / 'pdftoppm.exe'
if bundled.is_file():
    renderer = str(bundled)
checks = []
figure_specs = {
    'data_preprocess_diagnostics.pdf': (900, 350),
    'problem1_association_rates.pdf': (1200, 500),
    'problem1_centers_dumbbell.pdf': (1200, 600),
    'problem1_change_sensitivity.pdf': (1200, 600),
    'problem1_ternary_centers.pdf': (1200, 500),
    'problem1_paired_cases.pdf': (1200, 600),
}
with TemporaryDirectory() as temp:
    temp_dir = Path(temp)
    for name, (min_width, min_height) in figure_specs.items():
        path = ROOT / 'figures' / name
        if not path.is_file() or path.stat().st_size < 1024 or path.read_bytes()[:5] != b'%PDF-':
            raise AssertionError(f'Invalid PDF: {path}')
        prefix = temp_dir / path.stem
        completed = run([renderer, '-f', '1', '-l', '1', '-r', '180', '-png', str(path), str(prefix)], capture_output=True, text=True)
        if completed.returncode:
            raise RuntimeError(f'Could not render {name}: {completed.stderr.strip()}')
        image_path = next(temp_dir.glob(f'{path.stem}-*.png'), None)
        if image_path is None:
            raise AssertionError(f'No raster preview produced for {name}.')
        image = Image.open(image_path).convert('RGB')
        if image.width < min_width or image.height < min_height:
            raise AssertionError(f'Unexpected rendered size for {name}: {image.size}')
        nonwhite = ImageChops.difference(image, Image.new('RGB', image.size, 'white')).getbbox()
        if nonwhite is None:
            raise AssertionError(f'Rendered figure is blank: {name}')
        preview = ROOT / 'figures' / f'{path.stem}.png'
        if not preview.is_file():
            raise AssertionError(f'Missing 300 dpi PNG preview: {preview}')
        preview_image = Image.open(preview)
        if preview_image.width < min_width or preview_image.height < min_height:
            raise AssertionError(f'Unexpected PNG size for {preview.name}: {preview_image.size}')
        dpi = preview_image.info.get('dpi', (0, 0))
        if min(dpi) < 299:
            raise AssertionError(f'PNG preview is not 300 dpi: {preview.name}, dpi={dpi}')
        checks.append(f'{name}: rendered {image.width}x{image.height}, nonblank bbox={nonwhite}; {preview.name}: {preview_image.size} at {dpi[0]:.1f} dpi')

report = ROOT / 'data' / 'figure_render_check.txt'
report.write_text('PDF rendering check: PASS\n' + '\n'.join(checks) + '\n', encoding='utf-8')
print(report.read_text(encoding='utf-8'), end='')

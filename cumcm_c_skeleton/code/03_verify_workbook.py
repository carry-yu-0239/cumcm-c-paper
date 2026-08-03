"""XLSX structure and PDF signature checks only; no numerical processing."""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'data' / '\u9644\u4ef6_\u6570\u636e\u9884\u5904\u7406\u7ed3\u679c.xlsx'
REQUIRED = {'\u5904\u7406\u8bf4\u660e', '\u8868\u53551\u6e05\u6d17\u53ca\u989c\u8272\u586b\u8865', '\u8868\u53552\u6e05\u6d17\u8bb0\u5f55', '\u6709\u6548\u539f\u59cb\u6570\u636e', '\u95ed\u5408\u5f52\u4e00\u5316\u6570\u636e', '\u8fd1\u96f6\u66ff\u6362\u6570\u636e', 'CLR\u539f\u59cb\u503c', 'Z-CLR\u6570\u636e', '\u5f02\u5e38\u8bca\u65ad', '\u8fd1\u96f6\u53c2\u6570\u654f\u611f\u6027', 'U3_closed', 'U3_replace', 'U3_CLR', 'U3_ZCLR'}

if not BOOK.is_file(): raise FileNotFoundError(BOOK)
wb = load_workbook(BOOK, read_only=True, data_only=True)
missing = REQUIRED - set(wb.sheetnames)
if missing: raise AssertionError(f'Missing sheets: {sorted(missing)}')
if wb['Z-CLR\u6570\u636e'].max_column != 16: raise AssertionError('Z-CLR must have 2 identifiers plus 14 variables.')
if wb['U3_ZCLR'].max_column != 15: raise AssertionError('Unknown Z-CLR must have 1 identifier plus 14 variables.')
for name in ('data_preprocess_flow.pdf','data_preprocess_diagnostics.pdf'):
    path = ROOT / 'figures' / name
    if not path.is_file() or path.stat().st_size < 1024 or path.read_bytes()[:5] != b'%PDF-': raise AssertionError(f'Invalid PDF: {path}')
print('XLSX structure and PDF signatures: PASS')
